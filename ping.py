import pingparsing, json, datetime, random
import pandas as pd
from openpyxl import Workbook, load_workbook




if __name__ == '__main__':
    ADDRESS_LIST = ['1.1.1.1', '8.8.8.8', '8.8.4.4']
    rand = random.randrange(0, len(ADDRESS_LIST))
    ADDRESS = ADDRESS_LIST[rand]
    PACKETS = 10
    xlsx = 'ping.xlsx'

    date = datetime.datetime.now()
    date = date.strftime('%d/%m/%Y %H:%M')

    ping_parser = pingparsing.PingParsing()
    transmitter = pingparsing.PingTransmitter()
    transmitter.destination = ADDRESS
    transmitter.count = PACKETS
    result = transmitter.ping()
    json_result = json.dumps(ping_parser.parse(result).as_dict(), indent=4)
    dict_result = json.loads(json_result)
    #print(json_result)

    columns = ['avg_ping[ms]','packets_sent','packet_loss[%]','destination','date']

    for key, value in dict_result.items():
        if key == "packet_loss_rate":
            packet_loss = str(value)
        elif key == "packet_transmit":
            packets_sent = str(value)
        elif key == "destination":
            destination = str(value)
        elif key == "rtt_avg":
            avg_ping = str(value)



    dict_data = {'avg_ping[ms]': [avg_ping], 'packets_sent':[packets_sent],'packet_loss[%]':[packet_loss],'destination': [destination],'date':[date]}

    df = pd.DataFrame(dict_data)


    def append():
        try:
            old_df = pd.read_excel(xlsx, sheet_name='ping')
            pd.concat([df, old_df]).to_excel(xlsx, sheet_name='ping', index= False)
        except ValueError:
            wb = load_workbook(xlsx)
            wb.create_sheet(title="ping")
            print("Created 'ping' sheet")
            sheets = wb.sheetnames
            i = 0
            for sheet in sheets:

                if sheet !='ping': 
                    wb.remove(wb[sheet])
                    print(f"Removed '{sheet}' sheet in {xlsx}")
                i+=1

            wb.save(xlsx)
            append()


    try:
        try:
            append()
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            wb.save(xlsx)
            print("file created")
            append()
    except PermissionError:
        print("PermissionError please close your xlsx file")



    #print(df)







